import calendar
import io
import re
import zipfile
import xml.etree.ElementTree as XET
from copy import copy
from datetime import date, datetime
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font

st.title("Monthly Line Listing Application")

CELIX_PRODUCTS = [
    "abiraterone", "apixaban", "apremilast", "bexarotene", "brivaracetam", "clobazam",
    "clonazepam", "cyanocobalamin", "dabigatran", "dapagliflozin", "dimethyl fumarate",
    "edoxaban", "empagliflozin", "famotidine", "fesoterodine", "icatibant", "itraconazole",
    "linagliptin + metformin", "linagliptin", "metformin", "letrozole", "nintedanib",
    "pirfenidone", "raltegravir", "ranolazine", "rivaroxaban", "safinamide", "saxagliptin",
    "sitagliptin", "sacubitril + valsartan", "sacubritril + valsartan", "sacubitril",
    "sacubritril", "valsartan", "tamsulosin + solifenacin", "tapentadol", "ticagrelor",
    "tamsulosin", "solifenacin",
]

# Product-name variations that can appear in the line listing.
ALIASES = {
    "dabigatran": ["dabigatran etexilate"],
    "fesoterodine": ["fesoterodine fumarate"],
    "tamsulosin + solifenacin": ["tamsulosin/solifenacin", "solifenacin + tamsulosin", "vacit", "vecit"],
    "sacubitril + valsartan": ["sacubitril/valsartan", "sacubitril valsartan"],
    "sacubritril + valsartan": ["sacubritril/valsartan", "sacubritril valsartan"],
}


def norm(value):
    text = str(value or "").lower()
    text = re.sub(r"[^a-z0-9+]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def product_patterns():
    patterns = {}
    for product in CELIX_PRODUCTS:
        canonical = norm(product)
        patterns.setdefault(canonical, set()).add(canonical)
        for alias in ALIASES.get(product, []):
            patterns[canonical].add(norm(alias))
    # Combination products are checked first, preventing accidental single-ingredient matching.
    return {k: sorted(v, key=len, reverse=True) for k, v in patterns.items()}


PRODUCT_PATTERNS = product_patterns()


def exact_product_tokens(product_cell):
    # Product fields in the report are semicolon separated. Evaluate each product independently.
    return [norm(part) for part in re.split(r"[;\n]+", str(product_cell or "")) if norm(part)]


def matched_celix_products(product_cell):
    matches = []
    tokens = exact_product_tokens(product_cell)
    for canonical, aliases in sorted(PRODUCT_PATTERNS.items(), key=lambda item: len(item[0]), reverse=True):
        found = False
        for token in tokens:
            for alias in aliases:
                if token == alias or token.startswith(alias + " ") or re.search(rf"\b{re.escape(alias)}\b", token):
                    found = True
                    break
            if found:
                break
        if found and canonical not in matches:
            matches.append(canonical)
    # If a combination is present, do not also create separate sheets for its component words.
    if "tamsulosin + solifenacin" in matches:
        matches = [m for m in matches if m not in {"tamsulosin", "solifenacin"}]
    if "linagliptin + metformin" in matches:
        matches = [m for m in matches if m not in {"linagliptin", "metformin"}]
    if "sacubitril + valsartan" in matches or "sacubritril + valsartan" in matches:
        matches = [m for m in matches if m not in {"sacubitril", "sacubritril", "valsartan"}]
    return matches


def find_header_row(ws):
    for row in range(1, min(ws.max_row, 50) + 1):
        values = [norm(ws.cell(row, col).value) for col in range(1, min(ws.max_column, 20) + 1)]
        if "safety report id" in values and "product name" in values:
            return row
    raise ValueError("Could not find the header row containing Safety Report ID and Product Name.")


def find_column(ws, header_row, name):
    wanted = norm(name)
    for col in range(1, ws.max_column + 1):
        if norm(ws.cell(header_row, col).value) == wanted:
            return col
    raise ValueError(f"Required column not found: {name}")


def clone_style(source, target):
    if source.has_style:
        target._style = copy(source._style)
    if source.hyperlink:
        target._hyperlink = copy(source.hyperlink)
    if source.comment:
        target.comment = copy(source.comment)


def safe_sheet_name(product, used):
    base = re.sub(r"[\\/*?:\[\]]", " ", product.upper()).strip()[:31] or "PRODUCT"
    name = base
    counter = 2
    while name.lower() in used:
        suffix = f"_{counter}"
        name = base[:31 - len(suffix)] + suffix
        counter += 1
    used.add(name.lower())
    return name


def product_display(product):
    return product.upper()


def set_top_details(ws, product, start_date, end_date):
    display = product_display(product)
    replacements = {
        3: f"Project Name: CELIXP-GB-({display})({display})",
        4: f"Active Ingredient: {display}",
        5: f"Proprietory Name: {display}",
        6: f"Drug Name: {display}",
        7: f"Period: ADR Receipt Date: :{start_date.strftime('%d-%b-%y')} To {end_date.strftime('%d-%b-%y')}",
    }
    for row, value in replacements.items():
        ws.cell(row, 1).value = value



def remove_blank_top_rows(ws):
    """Remove fully blank leading rows while preserving values, styles and merged ranges."""
    blank_count = 0
    for row in range(1, ws.max_row + 1):
        if all(ws.cell(row, col).value in (None, "") for col in range(1, ws.max_column + 1)):
            blank_count += 1
        else:
            break
    if blank_count == 0:
        return
    merges = [
        (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
        for rng in ws.merged_cells.ranges
        if rng.max_row > blank_count
    ]
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    max_row, max_col = ws.max_row, ws.max_column
    for source_row in range(blank_count + 1, max_row + 1):
        target_row = source_row - blank_count
        for col in range(1, max_col + 1):
            source = ws.cell(source_row, col)
            target = ws.cell(target_row, col)
            target.value = source.value
            clone_style(source, target)
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    ws.delete_rows(max_row - blank_count + 1, blank_count)
    for min_row, min_col, max_row_merge, max_col_merge in merges:
        ws.merge_cells(
            start_row=max(1, min_row - blank_count),
            start_column=min_col,
            end_row=max_row_merge - blank_count,
            end_column=max_col_merge,
        )

def add_serial_number_column(ws, header_row):
    """Add Sl No only to the tabular area; leave the top report rows unchanged."""
    original_max_col = ws.max_column
    original_max_row = ws.max_row

    # Shift header and data cells one column to the right without inserting a full worksheet column.
    for row in range(original_max_row, header_row - 1, -1):
        for col in range(original_max_col, 0, -1):
            source = ws.cell(row, col)
            target = ws.cell(row, col + 1)
            target.value = source.value
            clone_style(source, target)
            source.value = None

    # Shift visible column widths for the table columns.
    for col in range(original_max_col, 0, -1):
        source_letter = get_column_letter(col)
        target_letter = get_column_letter(col + 1)
        ws.column_dimensions[target_letter].width = ws.column_dimensions[source_letter].width

    ws.cell(header_row, 1).value = "Sl No"
    clone_style(ws.cell(header_row, 2), ws.cell(header_row, 1))
    ws.column_dimensions["A"].width = 8


def normalize_xlsx_font_order(xlsx_bytes):
    """Preserve Excel-compatible font element order after openpyxl serialization."""
    source = io.BytesIO(xlsx_bytes)
    target = io.BytesIO()
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    order = {name: index for index, name in enumerate([
        "b", "i", "strike", "outline", "shadow", "condense", "extend",
        "sz", "color", "name", "family", "charset", "scheme", "u", "vertAlign"
    ])}
    with zipfile.ZipFile(source, "r") as zin, zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                root = XET.fromstring(data)
                fonts = root.find(f"{{{ns}}}fonts")
                if fonts is not None:
                    for font in fonts.findall(f"{{{ns}}}font"):
                        children = list(font)
                        children.sort(key=lambda child: order.get(child.tag.rsplit("}", 1)[-1], 999))
                        font[:] = children
                data = XET.tostring(root, encoding="utf-8", xml_declaration=False)
            zout.writestr(item, data)
    return target.getvalue()

def date_only_value(value):
    """Return ADR receipt value as a date only, without time."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%d-%b-%Y-%H:%M:%S", "%d-%b-%y-%H:%M:%S", "%d-%b-%Y", "%d-%b-%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return text


def semicolon_to_newline(value):
    """Display each semicolon-separated value on a separate line."""
    if not isinstance(value, str) or ";" not in value:
        return value
    parts = [part.strip() for part in value.split(";")]
    return "\n".join(part for part in parts if part)


def apply_output_cell_format(cell):
    """Wrap text and preserve the cell's existing alignment settings."""
    existing = cell.alignment
    cell.alignment = Alignment(
        horizontal=existing.horizontal,
        vertical=existing.vertical or "top",
        text_rotation=existing.text_rotation,
        wrap_text=True,
        shrink_to_fit=existing.shrink_to_fit,
        indent=existing.indent,
    )


def fit_product_sheet(ws, header_row):
    """Precisely fit columns while preserving one explicit line per LLT/PT/SOC entry."""
    multiline_headers = {"llt", "pt", "soc"}
    fixed_widths = {
        "sl no": 6.5,
        "safety report id": 20,
        "adr source":11,
        "adr receipt date": 15.0,
        "source country": 17,
        "case seriousness": 16.5,
        "case listedness": 16.0,
        "patient details": 21.0,
        "product name": 20.0,
        "event seriousness": 16.0,
        "event listedness": 15.0,
    }

    # Headers must always stay on one line and use a close-fit width.
    for col in range(1, ws.max_column + 1):
        header_cell = ws.cell(header_row, col)
        header_cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )
        header_cell.font = Font(
            name=header_cell.font.name,
            size=header_cell.font.size,
            bold=True,
            color="1F1F1F"
        )
        header_text = str(header_cell.value or "")
        existing = header_cell.alignment
        header_cell.alignment = Alignment(
            horizontal=existing.horizontal,
            vertical=existing.vertical or "center",
            text_rotation=existing.text_rotation,
            wrap_text=False,
            shrink_to_fit=False,
            indent=existing.indent,
        )
        header = norm(header_text)
        longest_explicit_line = len(header_text)
        for row in range(header_row + 1, ws.max_row + 1):
            value = str(ws.cell(row, col).value or "")
            longest_explicit_line = max(
                longest_explicit_line,
                max((len(line) for line in value.split("\n")), default=0),
            )

        if header in multiline_headers:
            # Width fits the longest individual term. Explicit new lines remain,
            # but no term should soft-wrap within its own line.
            width = min(max(longest_explicit_line + 0.5, len(header_text) + 0.5, 14), 72)
        elif header in fixed_widths:
            width = max(fixed_widths[header], len(header_text) + 0.5)
        else:
            width = min(max(longest_explicit_line + 0.5, len(header_text) + 0.5, 9), 38)
        ws.column_dimensions[get_column_letter(col)].width = width

    # LLT/PT/SOC require wrap_text=True only to display explicit newline characters.
    # Their widths are sized so each individual entry remains on one unwrapped line.
    for row in range(header_row + 1, ws.max_row + 1):
        required_lines = 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            header = norm(ws.cell(header_row, col).value)
            existing = cell.alignment
            explicit_lines = str(cell.value or "").split("\n") or [""]
            if header in multiline_headers:
                line_count = len(explicit_lines)
            else:
                width = max(int(ws.column_dimensions[get_column_letter(col)].width or 10), 1)
                line_count = sum(max(1, (len(line) + width - 1) // width) for line in explicit_lines)
            required_lines = max(required_lines, line_count)
            cell.alignment = Alignment(
                horizontal=existing.horizontal,
                vertical="top",
                text_rotation=existing.text_rotation,
                wrap_text=True,
                shrink_to_fit=False,
                indent=existing.indent,
            )
        ws.row_dimensions[row].height = min(max(18, required_lines * 15), 180)

    ws.row_dimensions[header_row].height = 20


def build_split_workbook(uploaded_bytes, selected_year, selected_month):
    workbook = load_workbook(io.BytesIO(uploaded_bytes))
    source = workbook[workbook.sheetnames[0]]
    # The source report may contain a remote Print Date cell in the otherwise blank first row.
    # Remove that auxiliary row so the report title starts immediately at row 1.
    for col in range(1, source.max_column + 1):
        source.cell(1, col).value = None
    remove_blank_top_rows(source)
    header_row = find_header_row(source)
    product_col = find_column(source, header_row, "Product Name")
    try:
        date_col = find_column(source, header_row, "ADR Receipt Date/Time")
    except ValueError:
        date_col = find_column(source, header_row, "ADR Receipt Date")
    case_listedness_col = find_column(source, header_row, "Case Listedness")
    event_listedness_col = find_column(source, header_row, "Event Listedness")

    first_day = date(selected_year, selected_month, 1)
    last_day = date(selected_year, selected_month, calendar.monthrange(selected_year, selected_month)[1])

    product_rows = {}
    for row in range(header_row + 1, source.max_row + 1):
        safety_id = source.cell(row, 1).value
        if safety_id in (None, ""):
            continue
        # Do not filter rows by ADR Receipt Date. The selected month controls only
        # the Period text at the top of each generated product sheet.
        for product in matched_celix_products(source.cell(row, product_col).value):
            product_rows.setdefault(product, []).append(row)

    template_name = source.title
    used_names = set()
    created = []
    for product in sorted(product_rows):
        ws = workbook.copy_worksheet(source)
        ws.title = safe_sheet_name(product, used_names)
        keep_rows = product_rows[product]
        # Remove existing data and rewrite matching rows from the original source sheet.
        if ws.max_row > header_row:
            ws.delete_rows(header_row + 1, ws.max_row - header_row)
        for output_index, source_row in enumerate(keep_rows, start=header_row + 1):
            for col in range(1, source.max_column + 1):
                src_cell = source.cell(source_row, col)
                copied_value = src_cell.value
                if col == date_col:
                    copied_value = date_only_value(copied_value)
                elif isinstance(copied_value, str):
                    copied_value = semicolon_to_newline(copied_value)
                dst_cell = ws.cell(output_index, col, copied_value)
                clone_style(src_cell, dst_cell)
                apply_output_cell_format(dst_cell)
                if col == date_col and isinstance(copied_value, (date, datetime)):
                    dst_cell.number_format = "dd-mmm-yyyy"
            ws.row_dimensions[output_index].height = source.row_dimensions[source_row].height

            # Ignore the supplied Case Listedness and derive it from Event Listedness.
            # Any Unexpected event makes the case Unexpected; otherwise it is Expected.
            event_listedness_value = str(source.cell(source_row, event_listedness_col).value or "")
            event_statuses = [
                norm(part) for part in re.split(r"[;\n]+", event_listedness_value) if norm(part)
            ]
            derived_case_listedness = (
                "Unexpected" if any(status == "unexpected" or status.endswith(" unexpected") for status in event_statuses)
                else "Expected"
            )
            ws.cell(output_index, case_listedness_col).value = derived_case_listedness
            apply_output_cell_format(ws.cell(output_index, case_listedness_col))

            # Keep all report columns, but show only the current Celix product in Product Name.
            ws.cell(output_index, product_col).value = product_display(product)
        set_top_details(ws, product, first_day, last_day)

        # Keep the report title left-aligned even though the title row is merged.
        title_cell = ws.cell(1, 1)
        title_alignment = title_cell.alignment
        title_cell.alignment = Alignment(
            horizontal="left",
            vertical=title_alignment.vertical or "center",
            text_rotation=title_alignment.text_rotation,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0,
        )

        add_serial_number_column(ws, header_row)
        output_date_col = date_col + 1
        ws.cell(header_row, output_date_col).value = "ADR Receipt Date"
        apply_output_cell_format(ws.cell(header_row, output_date_col))
        for index, row in enumerate(range(header_row + 1, ws.max_row + 1), start=1):
            ws.cell(row, 1).value = index
            clone_style(ws.cell(row, 2), ws.cell(row, 1))
            ws.cell(row, 1).alignment = copy(ws.cell(row, 2).alignment)
        fit_product_sheet(ws, header_row)
        ws.freeze_panes = None
        ws.auto_filter.ref = None
        ws.sheet_view.showGridLines = source.sheet_view.showGridLines
        created.append((product_display(product), len(keep_rows)))

    workbook.remove(workbook[template_name])
    if not created:
        raise ValueError("No Celix product rows were found in the uploaded report.")
    output = io.BytesIO()
    workbook.save(output)
    result_bytes = normalize_xlsx_font_order(output.getvalue())
    return result_bytes, created



def extract_pt_term(value):
    """Remove the trailing MedDRA numeric code from a PT display value."""
    return re.sub(r"\s*\(\s*\d+\s*\)\s*$", "", str(value or "")).strip()


def report_expectedness(value):
    text = norm(value)
    if text.endswith("unexpected") or " unexpected" in f" {text}":
        return "Unexpected"
    if text.endswith("expected") or " expected" in f" {text}":
        return "Expected"
    return ""


def listedness_index_from_repository():
    from listedness_service import dataframe_to_index
    data_dir = Path(__file__).resolve().parents[1] / "data"
    candidates = [data_dir / "Listedness_CX.xlsx", data_dir / "Listedness_CX.csv"]
    for path in candidates:
        if path.exists():
            if path.suffix.lower() == ".csv":
                import pandas as pd
                df = pd.read_csv(path)
            else:
                import pandas as pd
                df = pd.read_excel(path, engine="openpyxl")
            return dataframe_to_index(df), path.name
    raise FileNotFoundError("Listedness_CX.xlsx was not found in the data folder.")


def read_event_summary(uploaded_bytes):
    import pandas as pd
    workbook = load_workbook(io.BytesIO(uploaded_bytes), data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    header_row = find_header_row(ws)
    headers = [str(ws.cell(header_row, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    required = ["Safety Report ID", "Product Name", "PT", "Other Listedness"]
    positions = {}
    for name in required:
        target = norm(name)
        for col, header in enumerate(headers, start=1):
            if norm(header) == target:
                positions[name] = col
                break
        if name not in positions:
            raise ValueError(f"Required column not found: {name}")

    master, master_name = listedness_index_from_repository()
    checked, mismatches, missing = [], [], []
    for row in range(header_row + 1, ws.max_row + 1):
        safety_id = str(ws.cell(row, positions["Safety Report ID"]).value or "").strip()
        if not safety_id:
            continue
        product_text = ws.cell(row, positions["Product Name"]).value
        pt_raw = ws.cell(row, positions["PT"]).value
        pt = extract_pt_term(pt_raw)
        reported = report_expectedness(ws.cell(row, positions["Other Listedness"]).value)
        celix_products = matched_celix_products(product_text)
        for product in celix_products:
            key = (norm(product), norm(pt))
            stored = master.get(key)
            master_value = str((stored or {}).get("Expectedness", "")).strip()
            comment = str((stored or {}).get("Comment", "")).strip()
            if stored is None:
                status = "No exact match in master"
            elif not reported:
                status = "Source expectedness unavailable"
            elif norm(master_value) != norm(reported):
                status = "Mismatch"
            else:
                status = "Match"
            record = {
                "Safety Report ID": safety_id,
                "Celix Product": product.upper(),
                "PT": pt,
                "Report Expectedness": reported or "Not identified",
                "Master Expectedness": master_value or "Not found",
                "Status": status,
                "Master Comment": comment,
            }
            checked.append(record)
            if status == "Mismatch":
                mismatches.append(record)
            elif status == "No exact match in master":
                missing.append(record)
    return checked, mismatches, missing, master_name


def mismatch_workbook(records):
    import pandas as pd
    output = io.BytesIO()
    columns = ["Safety Report ID", "Celix Product", "PT", "Report Expectedness", "Master Expectedness", "Status", "Master Comment"]
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(records, columns=columns).to_excel(writer, index=False, sheet_name="Mismatch Pairs")
        ws = writer.book["Mismatch Pairs"]
        for cell in ws[1]:
            cell.font = copy(cell.font)
            cell.font = cell.font.copy(bold=True)
        widths = [24, 24, 36, 24, 24, 26, 42]
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width
    return output.getvalue()


tab_split, tab_check = st.tabs(["Product-wise Line Listing", "Event Listedness Check"])

with tab_split:
    st.subheader("Product-Wise Monthly Line Listing Splitter")
    st.caption("Upload one monthly line listing and download one workbook with a separate sheet for each Celix product.")
    uploaded = st.file_uploader("Upload monthly line listing", type=["xlsx"], key="line_listing_upload")
    current = date.today()
    c1, c2 = st.columns(2)
    with c1:
        selected_month = st.selectbox("Month", list(range(1, 13)), index=current.month - 1, format_func=lambda m: calendar.month_name[m])
    with c2:
        selected_year = st.number_input("Year", min_value=2000, max_value=2100, value=current.year, step=1)

    if uploaded and st.button("Generate product-wise workbook", type="primary"):
        try:
            result, summary = build_split_workbook(uploaded.getvalue(), int(selected_year), int(selected_month))
            st.success(f"Created {len(summary)} Celix product sheets.")
            st.dataframe({"Celix Product": [x[0] for x in summary], "Cases": [x[1] for x in summary]}, hide_index=True, use_container_width=True)
            filename = f"Celix_Product_Line_Listing_{int(selected_year)}_{int(selected_month):02d}.xlsx"
            st.download_button("Download segregated workbook", result, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as exc:
            st.error(f"Could not generate the workbook: {exc}")

with tab_check:
    import pandas as pd
    from listedness_service import github_upsert_rows, password_ok

    st.subheader("Event Summary Listedness Check")
    st.caption("Checks only Celix products and compares report expectedness against Active Ingredient + PT in Listedness_CX.xlsx.")

    if st.button("Clear uploaded file and displayed results", key="clear_event_listedness"):
        for key in [
            "event_listedness_results",
            "event_summary_upload",
            "listedness_batch_editor",
            "listedness_batch_password",
            "excluded_listedness_pairs",
        ]:
            st.session_state.pop(key, None)
        st.session_state["event_upload_version"] = st.session_state.get("event_upload_version", 0) + 1
        st.rerun()

    event_upload_version = st.session_state.get("event_upload_version", 0)
    event_file = st.file_uploader(
        "Upload Event Summary Report",
        type=["xlsx"],
        key=f"event_summary_upload_{event_upload_version}",
    )

    if event_file and st.button("Check listedness mismatches", type="primary"):
        try:
            checked, mismatches, missing, master_name = read_event_summary(event_file.getvalue())
            st.session_state["event_listedness_results"] = {
                "checked": checked,
                "mismatches": mismatches,
                "missing": missing,
                "master_name": master_name,
            }
        except Exception as exc:
            st.error(f"Could not check the Event Summary Report: {exc}")
            st.session_state.pop("event_listedness_results", None)

    results = st.session_state.get("event_listedness_results")
    if results:
        checked = results["checked"]
        mismatches = results["mismatches"]
        missing = results["missing"]
        master_name = results["master_name"]

        st.caption(f"Listedness master used: {master_name}")
        metric1, metric2, metric3 = st.columns(3)
        metric1.metric("Celix pairs checked", len(checked))
        metric2.metric("Mismatch pairs", len(mismatches))
        metric3.metric("Missing master pairs", len(missing))

        if mismatches:
            st.error(f"{len(mismatches)} listedness mismatch pair(s) found.")
            st.dataframe(mismatches, hide_index=True, use_container_width=True)
            st.download_button(
                "Download mismatch report",
                mismatch_workbook(mismatches),
                file_name="Celix_Listedness_Mismatch_Pairs.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.success("No listedness mismatches were found for Celix product pairs.")

        if missing:
            st.warning(f"{len(missing)} Active Ingredient + PT occurrence(s) were not found in the listedness master.")

        # Combine unique mismatch and missing pairs and retain all supporting Safety Report IDs.
        excluded_pairs = set(st.session_state.get("excluded_listedness_pairs", []))
        batch_pairs = {}

        def add_batch_occurrence(item, pair_type):
            pair_key = (norm(item["Celix Product"]), norm(item["PT"]))
            stable_key = f"{pair_key[0]}|||{pair_key[1]}"
            if stable_key in excluded_pairs:
                return
            safety_id = str(item.get("Safety Report ID", "") or "").strip()
            if pair_key not in batch_pairs:
                batch_pairs[pair_key] = {
                    "Delete": False,
                    "Update": True,
                    "Pair Type": pair_type,
                    "Safety Report ID": safety_id,
                    "Active Ingredients": item["Celix Product"],
                    "PT": item["PT"],
                    "Report Expectedness": item["Report Expectedness"],
                    "Current Master": item["Master Expectedness"] if pair_type == "Mismatch" else "Not found",
                    "New Expectedness": "",
                    "Comment": "" if str(item.get("Master Comment", "")).lower() == "nan" else str(item.get("Master Comment", "") or ""),
                    "_pair_key": stable_key,
                }
            elif safety_id:
                existing_ids = [value.strip() for value in batch_pairs[pair_key]["Safety Report ID"].split(";") if value.strip()]
                if safety_id not in existing_ids:
                    existing_ids.append(safety_id)
                    batch_pairs[pair_key]["Safety Report ID"] = "; ".join(existing_ids)

        for item in mismatches:
            add_batch_occurrence(item, "Mismatch")
        for item in missing:
            add_batch_occurrence(item, "Missing")

        if batch_pairs:
            st.markdown("#### Update missing and mismatch pairs")
            st.caption(
                "Safety Report ID shows the case source for verification. Select Delete to remove an incorrectly captured pair "
                "from this displayed batch, or manually select New Expectedness before updating GitHub."
            )
            display_columns = [
                "Delete", "Update", "Pair Type", "Safety Report ID", "Active Ingredients", "PT",
                "Report Expectedness", "Current Master", "New Expectedness", "Comment", "_pair_key"
            ]
            batch_df = pd.DataFrame(batch_pairs.values(), columns=display_columns)
            edited_batch = st.data_editor(
                batch_df,
                hide_index=True,
                use_container_width=True,
                key="listedness_batch_editor",
                column_config={
                    "Delete": st.column_config.CheckboxColumn("Delete", default=False),
                    "Update": st.column_config.CheckboxColumn("Update", default=True),
                    "Pair Type": st.column_config.TextColumn("Pair Type", disabled=True),
                    "Safety Report ID": st.column_config.TextColumn("Safety Report ID", disabled=True, width="large"),
                    "Active Ingredients": st.column_config.TextColumn("Active Ingredients", disabled=True),
                    "PT": st.column_config.TextColumn("PT", disabled=True, width="large"),
                    "Report Expectedness": st.column_config.TextColumn("Report Expectedness", disabled=True),
                    "Current Master": st.column_config.TextColumn("Current Master", disabled=True),
                    "New Expectedness": st.column_config.SelectboxColumn(
                        "New Expectedness", options=["Expected", "Unexpected"], required=False
                    ),
                    "Comment": st.column_config.TextColumn("Comment", width="large"),
                    "_pair_key": None,
                },
                disabled=[
                    "Pair Type", "Safety Report ID", "Active Ingredients", "PT",
                    "Report Expectedness", "Current Master", "_pair_key"
                ],
            )

            if st.button("Delete selected rows", key="delete_selected_listedness_rows"):
                selected_for_delete = edited_batch[edited_batch["Delete"] == True]
                if selected_for_delete.empty:
                    st.info("Select at least one Delete checkbox first.")
                else:
                    updated_exclusions = set(st.session_state.get("excluded_listedness_pairs", []))
                    updated_exclusions.update(selected_for_delete["_pair_key"].astype(str).tolist())
                    st.session_state["excluded_listedness_pairs"] = sorted(updated_exclusions)
                    st.session_state.pop("listedness_batch_editor", None)
                    st.rerun()

            batch_password = st.text_input(
                "Administrator password for all selected pairs",
                type="password",
                key="listedness_batch_password",
            )

            if st.button("Update all selected listedness pairs", key="submit_listedness_batch"):
                selected_rows = []
                skipped = 0
                for row in edited_batch.to_dict("records"):
                    if row.get("Delete") or not row.get("Update"):
                        continue
                    expectedness = str(row.get("New Expectedness", "") or "").strip()
                    if expectedness not in {"Expected", "Unexpected"}:
                        skipped += 1
                        continue
                    selected_rows.append({
                        "Active Ingredients": str(row.get("Active Ingredients", "")).strip(),
                        "PT": str(row.get("PT", "")).strip(),
                        "Expectedness": expectedness,
                        "Comment": str(row.get("Comment", "") or "").strip(),
                    })

                if not password_ok(batch_password):
                    st.error("Invalid administrator password.")
                elif not selected_rows:
                    st.error("Select at least one pair and choose New Expectedness.")
                else:
                    try:
                        outcome = github_upsert_rows(selected_rows)
                        st.success(
                            f"Listedness master updated in one GitHub commit. "
                            f"Added: {outcome['added']}; Updated: {outcome['updated']}; "
                            f"Commit: {outcome['sha'][:10]}"
                        )
                        if skipped:
                            st.info(f"Skipped {skipped} selected pair(s) because New Expectedness was blank.")
                        st.cache_data.clear()
                        st.session_state.pop("event_listedness_results", None)
                        st.session_state.pop("listedness_batch_editor", None)
                        st.session_state.pop("listedness_batch_password", None)
                        st.session_state.pop("excluded_listedness_pairs", None)
                    except Exception as exc:
                        st.error(f"GitHub batch update failed: {exc}")
