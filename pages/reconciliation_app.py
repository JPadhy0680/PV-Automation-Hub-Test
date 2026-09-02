import io
import re
from datetime import datetime, date
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter



def norm_header(value):
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def norm_id(value):
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def clean_text(value):
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def parse_date(value):
    if value is None or pd.isna(value) or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    formats = [
        "%d-%b-%Y-%H:%M:%S", "%d-%b-%y-%H:%M:%S", "%d-%b-%Y", "%d-%b-%y",
        "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
    return None if pd.isna(parsed) else parsed.date()


def display_date(value):
    parsed = parse_date(value)
    return parsed.strftime("%d-%b-%Y") if parsed else clean_text(value)


def norm_seriousness(value):
    text = norm_header(value)
    if text in {"non serious", "nonserious"} or "non serious" in text:
        return "Non-serious"
    if text == "serious" or ("serious" in text and "non" not in text):
        return "Serious"
    return clean_text(value)


def split_external_ids(value):
    text = clean_text(value)
    if not text:
        return set()
    return {norm_id(part) for part in re.split(r"[;,\n|]+", text) if norm_id(part)}


def reference_matches_external(reference_id, external_value):
    """Match exact IDs and report-truncated IDs using a conservative shared prefix."""
    reference = norm_id(reference_id)
    if not reference:
        return False
    for external in split_external_ids(external_value):
        if reference == external:
            return True
        # Reports frequently truncate long MHRA IDs or add a five-letter workflow suffix.
        shorter, longer = sorted([reference, external], key=len)
        if len(shorter) >= 20 and longer.startswith(shorter):
            return True
        reference_without_suffix = re.sub(r"-[A-Z]{5}$", "", reference)
        if reference_without_suffix == external:
            return True
        shorter, longer = sorted([reference_without_suffix, external], key=len)
        if len(shorter) >= 20 and longer.startswith(shorter):
            return True
    return False


def detect_header_row(ws, required_alias_groups):
    for row in range(1, min(ws.max_row, 60) + 1):
        headers = {norm_header(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        if all(any(norm_header(alias) in headers for alias in aliases) for aliases in required_alias_groups):
            return row
    raise ValueError("Could not identify the report header row.")


def read_report(file_bytes, report_type):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    if report_type == "tracker":
        required = [["Receipt Date"], ["Referance ID", "Reference ID"], ["Safety Repoprt ID", "Safety Report ID"]]
    else:
        required = [["Safety Report ID"], ["ADR Receipt Date/Time"], ["Case Seriousness"], ["External ID"]]
    header_row = detect_header_row(ws, required)
    headers = [clean_text(ws.cell(header_row, col).value) or f"Column {col}" for col in range(1, ws.max_column + 1)]
    rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        if any(value not in (None, "") for value in values):
            rows.append(values)
    return pd.DataFrame(rows, columns=headers)


def find_col(df, aliases, required=True):
    available = {norm_header(col): col for col in df.columns}
    for alias in aliases:
        if norm_header(alias) in available:
            return available[norm_header(alias)]
    if required:
        raise ValueError(f"Required column not found: {' / '.join(aliases)}")
    return None


def reconcile(tracker_df, safety_df):
    tc = {
        "ack": find_col(tracker_df, ["ACK No"], False),
        "receipt": find_col(tracker_df, ["Receipt Date"]),
        "source": find_col(tracker_df, ["Source"], False),
        "reference": find_col(tracker_df, ["Referance ID", "Reference ID"]),
        "ird": find_col(tracker_df, ["IRD"], False),
        "product": find_col(tracker_df, ["Suspect Product"], False),
        "validity": find_col(tracker_df, ["Validity"], False),
        "seriousness": find_col(tracker_df, ["Seriousness"]),
        "safety": find_col(tracker_df, ["Safety Repoprt ID", "Safety Report ID"]),
    }
    sc = {
        "safety": find_col(safety_df, ["Safety Report ID"]),
        "adr": find_col(safety_df, ["ADR Receipt Date/Time"]),
        "pv": find_col(safety_df, ["Pv Received Date/Time", "PV Received Date/Time"], False),
        "seriousness": find_col(safety_df, ["Case Seriousness"]),
        "external": find_col(safety_df, ["External ID"]),
    }

    safety_by_id = {}
    for idx, row in safety_df.iterrows():
        sid = norm_id(row.get(sc["safety"]))
        if sid:
            safety_by_id.setdefault(sid, []).append((idx, row))

    used_safety_rows = set()
    results = []
    for _, trow in tracker_df.iterrows():
        tracker_sid = norm_id(trow.get(tc["safety"]))
        tracker_ref = norm_id(trow.get(tc["reference"]))
        candidates = safety_by_id.get(tracker_sid, []) if tracker_sid else []
        match_method = "Safety Report ID" if candidates else ""

        # Safety Report ID is authoritative when present and must match exactly,
        # including the final version suffix such as -01 or -02.
        # Reference ID fallback is allowed only when the tracker Safety Report ID is blank.
        if not tracker_sid and tracker_ref:
            for sidx, srow in safety_df.iterrows():
                if reference_matches_external(tracker_ref, srow.get(sc["external"])):
                    candidates.append((sidx, srow))
            if candidates:
                match_method = "Reference ID / External ID"

        if not candidates:
            results.append({
                "Status": "Missing in Safety Report",
                "Match Method": "",
                "ACK No": clean_text(trow.get(tc["ack"])) if tc["ack"] else "",
                "Tracker Safety Report ID": clean_text(trow.get(tc["safety"])),
                "Safety System ID": "",
                "Reference ID": clean_text(trow.get(tc["reference"])),
                "External ID": "",
                "Reference ID Check": "Not checked",
                "Tracker Receipt Date": display_date(trow.get(tc["receipt"])),
                "ADR Receipt Date": "",
                "Receipt Date vs PV Check": "Not checked",
                "IRD vs ADR Check": "Not checked",
                "Tracker Seriousness": norm_seriousness(trow.get(tc["seriousness"])),
                "Safety Seriousness": "",
                "Seriousness Check": "Not checked",
                "Source": clean_text(trow.get(tc["source"])) if tc["source"] else "",
                "IRD": display_date(trow.get(tc["ird"])) if tc["ird"] else "",
                "PV Received Date/Time": "",
                "Suspect Product": clean_text(trow.get(tc["product"])) if tc["product"] else "",
                "Validity": clean_text(trow.get(tc["validity"])) if tc["validity"] else "",
                "Mismatch Details": (
                    "Exact Safety Report ID was not found in the safety-system report"
                    if tracker_sid else
                    "Reference ID was not found in the safety-system report"
                ),
            })
            continue

        # Prefer exact SID candidate if multiple fallback records are found.
        sidx, srow = candidates[0]
        used_safety_rows.add(sidx)
        safety_sid = clean_text(srow.get(sc["safety"]))
        reference_match = reference_matches_external(tracker_ref, srow.get(sc["external"]))
        receipt_match = parse_date(trow.get(tc["receipt"])) == parse_date(srow.get(sc["pv"])) if sc["pv"] else False
        ird_match = parse_date(trow.get(tc["ird"])) == parse_date(srow.get(sc["adr"])) if tc["ird"] else True
        seriousness_match = norm_header(norm_seriousness(trow.get(tc["seriousness"]))) == norm_header(norm_seriousness(srow.get(sc["seriousness"])))
        sid_match = bool(tracker_sid and tracker_sid == norm_id(safety_sid))

        issues = []
        if not sid_match:
            issues.append("Safety Report ID mismatch")
        if not reference_match:
            issues.append("Reference ID not found in External ID")
        if not receipt_match:
            issues.append("Receipt Date differs from PV Received Date")
        if not ird_match:
            issues.append("IRD differs from ADR Receipt Date")
        if not seriousness_match:
            issues.append("Seriousness mismatch")
        status = "Match" if not issues else "Mismatch"

        results.append({
            "Status": status,
            "Match Method": match_method,
            "ACK No": clean_text(trow.get(tc["ack"])) if tc["ack"] else "",
            "Tracker Safety Report ID": clean_text(trow.get(tc["safety"])),
            "Safety System ID": safety_sid,
            "Reference ID": clean_text(trow.get(tc["reference"])),
            "External ID": clean_text(srow.get(sc["external"])),
            "Reference ID Check": "Match" if reference_match else "Mismatch",
            "Tracker Receipt Date": display_date(trow.get(tc["receipt"])),
            "ADR Receipt Date": display_date(srow.get(sc["adr"])),
            "PV Received Date/Time": clean_text(srow.get(sc["pv"])) if sc["pv"] else "",
            "Receipt Date vs PV Check": "Match" if receipt_match else "Mismatch",
            "IRD vs ADR Check": "Match" if ird_match else "Mismatch",
            "Tracker Seriousness": norm_seriousness(trow.get(tc["seriousness"])),
            "Safety Seriousness": norm_seriousness(srow.get(sc["seriousness"])),
            "Seriousness Check": "Match" if seriousness_match else "Mismatch",
            "Source": clean_text(trow.get(tc["source"])) if tc["source"] else "",
            "IRD": display_date(trow.get(tc["ird"])) if tc["ird"] else "",
            "Suspect Product": clean_text(trow.get(tc["product"])) if tc["product"] else "",
            "Validity": clean_text(trow.get(tc["validity"])) if tc["validity"] else "",
            "Mismatch Details": "; ".join(issues),
        })

    # Safety-system records absent from the tracker.
    for sidx, srow in safety_df.iterrows():
        if sidx in used_safety_rows:
            continue
        results.append({
            "Status": "Missing in Tracker",
            "Match Method": "",
            "ACK No": "",
            "Tracker Safety Report ID": "",
            "Safety System ID": clean_text(srow.get(sc["safety"])),
            "Reference ID": "",
            "External ID": clean_text(srow.get(sc["external"])),
            "Reference ID Check": "Not checked",
            "Tracker Receipt Date": "",
            "ADR Receipt Date": display_date(srow.get(sc["adr"])),
            "Receipt Date vs PV Check": "Not checked",
                "IRD vs ADR Check": "Not checked",
            "Tracker Seriousness": "",
            "Safety Seriousness": norm_seriousness(srow.get(sc["seriousness"])),
            "Seriousness Check": "Not checked",
            "Source": "",
            "IRD": "",
            "PV Received Date/Time": clean_text(srow.get(sc["pv"])) if sc["pv"] else "",
            "Suspect Product": "",
            "Validity": "",
            "Mismatch Details": "Safety-system record was not found in the tracker",
        })
    return pd.DataFrame(results)


def output_workbook(result_df):
    output = io.BytesIO()
    exceptions = result_df[result_df["Status"] != "Match"].copy()
    matches = result_df[result_df["Status"] == "Match"].copy()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        exceptions.to_excel(writer, index=False, sheet_name="Exceptions")
        matches.to_excel(writer, index=False, sheet_name="Matched")
        result_df.to_excel(writer, index=False, sheet_name="All Results")
        for ws in writer.book.worksheets:
            ws.freeze_panes = None
            ws.auto_filter.ref = None
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            for row in ws.iter_rows(min_row=2):
                status = row[0].value
                if status == "Mismatch":
                    fill = PatternFill("solid", fgColor="FCE4D6")
                elif status in {"Missing in Safety Report", "Missing in Tracker"}:
                    fill = PatternFill("solid", fgColor="FFF2CC")
                else:
                    fill = None
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if fill:
                        cell.fill = fill
            for col in range(1, ws.max_column + 1):
                values = [clean_text(ws.cell(row, col).value) for row in range(1, min(ws.max_row, 100) + 1)]
                ws.column_dimensions[get_column_letter(col)].width = min(max(max((len(v) for v in values), default=10) + 2, 12), 45)
    return output.getvalue()


st.title("Monthly Reconciliation and Summary")
tab_reconciliation, tab_summary = st.tabs(["Case Reconciliation", "Data Summary Generator"])

with tab_reconciliation:
    st.header("Monthly Case Reconciliation")
    st.caption("Upload the tracker and safety-system reports to identify missing cases and field-level mismatches.")
    if st.button("Clear uploads and results", key="recon_clear"):
        for key in ["recon_results", "tracker_upload", "safety_upload"]:
            st.session_state.pop(key, None)
        st.session_state["recon_version"] = st.session_state.get("recon_version", 0) + 1
        st.rerun()

    version = st.session_state.get("recon_version", 0)
    left, right = st.columns(2)
    with left:
        tracker_file = st.file_uploader("Upload Tracker Report", type=["xlsx"], key=f"tracker_upload_{version}")
    with right:
        safety_file = st.file_uploader("Upload Safety-System Report", type=["xlsx"], key=f"safety_upload_{version}")

    if tracker_file and safety_file and st.button("Run reconciliation", type="primary"):
        try:
            tracker = read_report(tracker_file.getvalue(), "tracker")
            safety = read_report(safety_file.getvalue(), "safety")
            st.session_state["recon_results"] = reconcile(tracker, safety)
        except Exception as exc:
            st.error(f"Reconciliation failed: {exc}")

    result = st.session_state.get("recon_results")
    if isinstance(result, pd.DataFrame):
        match_count = int((result["Status"] == "Match").sum())
        mismatch_count = int((result["Status"] == "Mismatch").sum())
        missing_safety = int((result["Status"] == "Missing in Safety Report").sum())
        missing_tracker = int((result["Status"] == "Missing in Tracker").sum())
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Matched", match_count)
        c2.metric("Mismatched", mismatch_count)
        c3.metric("Missing in Safety", missing_safety)
        c4.metric("Missing in Tracker", missing_tracker)

        exceptions = result[result["Status"] != "Match"]
        if exceptions.empty:
            st.success("Reconciliation completed. No mismatches or missing cases were found.")
        else:
            st.error(f"{len(exceptions)} reconciliation exception(s) found.")
            st.dataframe(exceptions, hide_index=True, use_container_width=True)

        with st.expander("Show matched records"):
            st.dataframe(result[result["Status"] == "Match"], hide_index=True, use_container_width=True)

        st.download_button(
            "Download reconciliation workbook",
            output_workbook(result),
            file_name="Celix_Monthly_Reconciliation.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

with tab_summary:
    st.header("Data Summary Generator")
    st.caption("Upload the tracker Excel file to generate a date-wise and source-wise summary.")

    if st.button("Clear summary upload and results", key="summary_clear"):
        st.session_state.pop("summary_result", None)
        st.session_state["summary_version"] = st.session_state.get("summary_version", 0) + 1
        st.rerun()

    summary_version = st.session_state.get("summary_version", 0)
    summary_file = st.file_uploader("Upload Excel File", type=["xlsx"], key=f"summary_upload_{summary_version}")

    if summary_file and st.button("Generate summary", type="primary", key="generate_summary"):
        try:
            df = pd.read_excel(summary_file, engine="openpyxl")
            if df.empty:
                raise ValueError("The uploaded Excel file is empty.")
            df.columns = df.columns.astype(str).str.strip().str.lower()
            possible_dates = ["download date", "receipt date", "downloaded date", "date of receipt", "receiptdate", "download_date", "receipt_date"]
            date_col = next((c for c in df.columns if c in possible_dates), None)
            if not date_col:
                raise ValueError("No valid Receipt Date or Download Date column was found.")
            df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
            invalid = int(df[date_col].isna().sum())
            df = df.dropna(subset=[date_col]).sort_values(date_col)
            source_col = "source" if "source" in df.columns else None
            if not source_col:
                df["source"] = "UNKNOWN"; source_col = "source"
            df[source_col] = df[source_col].astype(str).str.strip().str.upper().replace(["", "NAN", "NONE"], "UNKNOWN")
            validity_col = "validity" if "validity" in df.columns else None
            rows=[]
            pairs=df[[date_col,source_col]].drop_duplicates().sort_values([date_col,source_col])
            for _,pair in pairs.iterrows():
                current=pair[date_col]; source=pair[source_col]
                subset=df[(df[date_col]==current)&(df[source_col]==source)]
                no_report=subset.astype(str).apply(lambda c:c.str.contains("No Report Received",case=False,na=False)).any().any()
                if source=="MHRA":
                    previous=pairs[(pairs[source_col]=="MHRA")&(pairs[date_col]<current)][date_col]
                    if previous.empty:
                        from_date=current-pd.Timedelta(days=3 if current.day_name()=="Monday" else 1)
                    else: from_date=previous.max()
                    to_date=current-pd.Timedelta(days=1)
                elif source in ["ADIS","NA"]:
                    from_date=current-pd.Timedelta(days=current.weekday()); to_date=current
                else: from_date=to_date=None
                if no_report: total=valid=non_valid="No Report Received"
                else:
                    total=len(subset)
                    if validity_col:
                        v=subset[validity_col].astype(str).str.strip().str.lower()
                        valid=int((v=="valid").sum()); non_valid=int((v=="non-valid").sum())
                    else: valid=non_valid=0
                rows.append({"Receipt Date":current,"Source":source,"From":from_date,"To":to_date,"Total Number":total,"Valid":valid,"Non-Valid":non_valid})
            summary=pd.DataFrame(rows)
            for col in ["Receipt Date","From","To"]:
                summary[col]=pd.to_datetime(summary[col],errors="coerce").dt.strftime("%d-%b-%y")
            st.session_state["summary_result"]={"data":summary.fillna(""),"invalid":invalid}
        except Exception as exc: st.error(f"Summary generation failed: {exc}")

    saved=st.session_state.get("summary_result")
    if saved:
        if saved["invalid"]: st.warning(f"Ignored {saved['invalid']} invalid or blank date row(s).")
        summary=saved["data"]
        st.dataframe(summary,use_container_width=True,hide_index=True)
        output=io.BytesIO()
        with pd.ExcelWriter(output,engine="openpyxl") as writer:
            summary.to_excel(writer,index=False,sheet_name="Summary")
            ws=writer.book["Summary"]; ws.freeze_panes=None; ws.auto_filter.ref=None
            for cell in ws[1]:
                cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="1F4E78")
            for col in range(1,ws.max_column+1): ws.column_dimensions[get_column_letter(col)].width=18
        st.download_button("Download Summary Excel",output.getvalue(),file_name="Summary_Output.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
